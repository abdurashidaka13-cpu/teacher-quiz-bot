import io
import random
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# O'zbekcha kirill harflarini lotinchaga transliteratsiya qilish lug'ati
CYRILLIC_TO_LATIN = {
    "а": "a",
    "б": "b",
    "в": "v",
    "г": "g",
    "д": "d",
    "е": "e",
    "ё": "yo",
    "ж": "j",
    "з": "z",
    "и": "i",
    "й": "y",
    "к": "k",
    "л": "l",
    "м": "m",
    "н": "n",
    "о": "o",
    "п": "p",
    "р": "r",
    "с": "s",
    "т": "t",
    "у": "u",
    "ф": "f",
    "х": "x",
    "ц": "ts",
    "ч": "ch",
    "ш": "sh",
    "щ": "sh",
    "ъ": "",
    "ы": "i",
    "ь": "",
    "э": "e",
    "ю": "yu",
    "я": "ya",
    "ў": "o'",
    "қ": "q",
    "ғ": "g'",
    "ҳ": "h",
    "А": "A",
    "Б": "B",
    "В": "V",
    "Г": "G",
    "Д": "D",
    "Е": "E",
    "Ё": "Yo",
    "Ж": "J",
    "З": "Z",
    "И": "I",
    "Й": "Y",
    "К": "K",
    "Л": "L",
    "М": "M",
    "Н": "N",
    "О": "O",
    "П": "P",
    "Р": "R",
    "С": "S",
    "Т": "T",
    "У": "U",
    "Ф": "F",
    "Х": "X",
    "Ц": "Ts",
    "Ч": "Ch",
    "Ш": "Sh",
    "Щ": "Sh",
    "Ъ": "",
    "Ы": "I",
    "Ь": "",
    "Э": "E",
    "Ю": "Yu",
    "Я": "Ya",
    "Ў": "O'",
    "Қ": "Q",
    "Ғ": "G'",
    "Ҳ": "H",
}


def transliterate_uzbek(text: str) -> str:
    """
    O'zbek ismlarini kirillchadan lotinchaga o'tkazadi
    va maxsus belgilarni tozalaydi.
    """
    res = []
    for char in text:
        res.append(CYRILLIC_TO_LATIN.get(char, char))
    latin_text = "".join(res)

    # Faqat harflar, raqamlar va pastki chiziqni qoldirish
    latin_text = latin_text.replace("'", "").replace("`", "")
    latin_text = re.sub(r"[^\w\s-]", "", latin_text)
    latin_text = re.sub(r"[\s-]+", "_", latin_text)
    return latin_text.lower().strip("_")


def generate_credentials(names: list) -> list:
    """
    O'quvchilar ro'yxatidan login va 5 xonali raqamli parol hosil qiladi.
    Dublikat ismlar bo'lsa, ali_valiev_1 kabi loginlar yaratadi.
    """
    students_data = []
    existing_logins = set()

    for idx, name in enumerate(names):
        name_clean = name.strip()
        if not name_clean:
            continue

        # Login hosil qilish
        base_login = transliterate_uzbek(name_clean)
        if not base_login:
            base_login = f"student_{idx + 1}"

        login = base_login
        counter = 1
        # Dublikat loginlarni hal qilish
        while login in existing_logins:
            login = f"{base_login}_{counter}"
            counter += 1

        existing_logins.add(login)

        # 5 xonali raqamli parol yaratish
        password = str(random.randint(10000, 99999))

        students_data.append({"full_name": name_clean, "login": login, "password": password})

    return students_data


def parse_students_excel(file_stream) -> dict:
    """
    O'quvchilar Excel jadvalini o'qib, ularga login-parol berib qaytaradi.
    """
    try:
        df = pd.read_excel(file_stream)
    except Exception as e:
        return {"error": f"Excel faylini o'qib bo'lmadi: {e}"}

    # Ustun nomi tekshiruvi
    # Katta-kichik harflar va chetki probellarni tozalab tekshiramiz
    cleaned_columns = [str(col).strip().lower() for col in df.columns]

    if "ism familiya" not in cleaned_columns:
        return {"error": "Jadvalda 'Ism Familiya' ustuni topilmadi! Iltimos, shablonni tekshiring."}

    # Haqiqiy ustun indeksini topish
    col_idx = cleaned_columns.index("ism familiya")
    col_name = df.columns[col_idx]

    # Bo'sh bo'lmagan ismlarni olish
    names = df[col_name].dropna().astype(str).tolist()
    names = [n.strip() for n in names if n.strip()]

    if not names:
        return {"error": "Excel jadvalida o'quvchilar ismlari topilmadi!"}

    students_credentials = generate_credentials(names)
    return {"students": students_credentials}


def generate_results_excel(quiz_title: str, students_list: list, questions_count: int, attempts_map: dict) -> bytes:
    """
    Imtihon natijalari yozilgan 2 varoqli baholash matritsasini yaratadi.
    attempts_map: {
        student_id: {
            'score': int,
            'completed': bool,
            'answers': { question_id (asl tartib): bool }
        }
    }
    """
    wb = Workbook()

    # 1-Varoq: Umumiy Natijalar (Sheet 1)
    ws1 = wb.active
    ws1.title = "Umumiy Natijalar"
    ws1.views.sheetView[0].showGridLines = True

    # Stil to'plamlari
    font_title = Font(name="Calibri", size=14, bold=True, color="1F497D")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=11)
    font_bold = Font(name="Calibri", size=11, bold=True)

    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")

    border_thin = Side(border_style="thin", color="D9D9D9")
    border_double = Side(border_style="double", color="333333")
    border_thick = Side(border_style="medium", color="1F497D")

    border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    border_footer = Border(top=border_thin, bottom=border_double)

    # 1-Varoq Sarlavhalari
    ws1["A1"] = f"Test: {quiz_title}"
    ws1["A1"].font = font_title

    headers1 = ["T/r", "Ism Familiya", "Jami savollar", "To'g'ri topilgan", "Foiz (%)"]
    for col_idx, header in enumerate(headers1, 1):
        cell = ws1.cell(row=3, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 1-Varoq Ma'lumotlarini to'ldirish
    row_num = 4
    for idx, student in enumerate(students_list, 1):
        att = attempts_map.get(student["id"])
        score = att["score"] if att else 0
        percentage = round((score / questions_count) * 100, 1) if questions_count > 0 else 0

        # Agar student qatnashmagan bo'lsa
        if not att or not att["completed"]:
            percentage_val = "Qatnashmadi"
            score_val = "-"
        else:
            percentage_val = f"{percentage}%"
            score_val = score

        ws1.cell(row=row_num, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_num, column=2, value=student["full_name"])
        ws1.cell(row=row_num, column=3, value=questions_count).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_num, column=4, value=score_val).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_num, column=5, value=percentage_val).alignment = Alignment(horizontal="center")

        # Stil qo'llash
        for col_idx in range(1, 6):
            cell = ws1.cell(row=row_num, column=col_idx)
            cell.font = font_data
            cell.border = border_cell
            if row_num % 2 == 0:
                cell.fill = fill_zebra

        row_num += 1

    # Ustunlar kengligini sozlash
    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # -------------------------------------------------------------
    # 2-Varoq: Baholash Matritsasi (Sheet 2)
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Baholash Matritsasi")
    ws2.views.sheetView[0].showGridLines = True

    # 2-Varoq Sarlavhalari
    ws2["A1"] = f"Baholash Matritsasi: {quiz_title}"
    ws2["A1"].font = font_title

    headers2 = ["T/r", "Familiyasi Ismi"]
    for q_idx in range(1, questions_count + 1):
        headers2.append(f"S-{q_idx}")
    headers2.extend(["Jami to'g'ri", "Foiz (%)"])

    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Matritsani o'quvchilar natijalari bilan to'ldirish
    start_row = 4
    student_count = len(students_list)
    active_students_rows = []

    for idx, student in enumerate(students_list, 1):
        curr_row = start_row + idx - 1
        ws2.cell(row=curr_row, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws2.cell(row=curr_row, column=2, value=student["full_name"])

        att = attempts_map.get(student["id"])
        is_active = att and att["completed"]

        # Savollar bo'yicha 1 yoki 0 to'ldirish
        for q_idx in range(1, questions_count + 1):
            col_pos = 2 + q_idx
            if is_active:
                # O'quvchi javobi asl tartibdagi question_id (q_idx) bo'yicha to'g'rimi?
                is_correct = att["answers"].get(q_idx)
                val = 1 if is_correct else 0
                ws2.cell(row=curr_row, column=col_pos, value=val).alignment = Alignment(horizontal="center")
            else:
                ws2.cell(row=curr_row, column=col_pos, value="").alignment = Alignment(horizontal="center")

        # Summa va Foiz formulalari
        first_q_letter = get_column_letter(3)
        last_q_letter = get_column_letter(2 + questions_count)

        if is_active:
            active_students_rows.append(curr_row)
            # Excel SUM formulasi
            ws2.cell(
                row=curr_row,
                column=3 + questions_count,
                value=f"=SUM({first_q_letter}{curr_row}:{last_q_letter}{curr_row})",
            ).alignment = Alignment(horizontal="center")

            # Excel FOIZ formulasi (Jami to'g'ri / jami savollar)
            sum_cell_letter = get_column_letter(3 + questions_count)
            ws2.cell(
                row=curr_row,
                column=4 + questions_count,
                value=f"={sum_cell_letter}{curr_row}/{questions_count}",
            ).alignment = Alignment(horizontal="center")
            ws2.cell(row=curr_row, column=4 + questions_count).number_format = "0.0%"
        else:
            # Qatnashmagan bo'lsa
            ws2.cell(row=curr_row, column=3 + questions_count, value="-").alignment = Alignment(
                horizontal="center"
            )
            ws2.cell(row=curr_row, column=4 + questions_count, value="Qatnashmadi").alignment = Alignment(
                horizontal="center"
            )

        # Matritsa kataklariga border va ranglar berish
        for col_pos in range(1, 5 + questions_count):
            cell = ws2.cell(row=curr_row, column=col_pos)
            cell.font = font_data
            cell.border = border_cell
            if curr_row % 2 == 0:
                cell.fill = fill_zebra

    # Eng pastki footer qatori: "Jami to'g'ri javoblar" (savollar bo'yicha yig'indi)
    footer_row = start_row + student_count
    ws2.cell(row=footer_row, column=1, value="").border = border_footer
    ws2.cell(row=footer_row, column=2, value="Jami to'g'ri javoblar:").font = font_bold
    ws2.cell(row=footer_row, column=2).border = border_footer

    first_stud_row = start_row
    last_stud_row = start_row + student_count - 1

    # Har bir savol ustuni bo'yicha 1 larni yig'ish formulasi
    for q_idx in range(1, questions_count + 1):
        col_pos = 2 + q_idx
        col_letter = get_column_letter(col_pos)
        cell = ws2.cell(
            row=footer_row,
            column=col_pos,
            value=f"=SUM({col_letter}{first_stud_row}:{col_letter}{last_stud_row})",
        )
        cell.font = font_bold
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_footer

    # O'ng burchakdagi jami to'g'ri va o'rtacha foiz formulalari
    sum_col_letter = get_column_letter(3 + questions_count)
    ws2.cell(
        row=footer_row,
        column=3 + questions_count,
        value=f"=SUM({sum_col_letter}{first_stud_row}:{sum_col_letter}{last_stud_row})",
    ).font = font_bold
    ws2.cell(row=footer_row, column=3 + questions_count).alignment = Alignment(horizontal="center")
    ws2.cell(row=footer_row, column=3 + questions_count).border = border_footer

    # Sinf O'rtacha Foizi (Average Class Percentage)
    pct_col_letter = get_column_letter(4 + questions_count)
    # AVERAGE funksiyasi Excelda Qatnashmadi (matnli) kataklarni hisobga olmaydi, bu pedagogik jihatdan to'g'ri.
    ws2.cell(
        row=footer_row,
        column=4 + questions_count,
        value=f"=AVERAGE({pct_col_letter}{first_stud_row}:{pct_col_letter}{last_stud_row})",
    ).font = font_bold
    ws2.cell(row=footer_row, column=4 + questions_count).alignment = Alignment(horizontal="center")
    ws2.cell(row=footer_row, column=4 + questions_count).number_format = "0.0%"
    ws2.cell(row=footer_row, column=4 + questions_count).border = border_footer

    # -------------------------------------------------------------
    # Bo'shliqlar Tahlili (Gap Analysis - Matritsa ostida 3 qator bo'sh joydan keyin)
    # -------------------------------------------------------------
    gap_row = footer_row + 3
    ws2.cell(
        row=gap_row,
        column=2,
        value="📌 Sinf bo'yicha aniqlangan bo'shliqlar (O'zlashtirilmagan savollar diagnostikasi):",
    ).font = Font(name="Calibri", size=11, bold=True, color="FF0000")

    # Dasturiy tarzda eng qiyin savollarni hisoblash
    q_stats = []
    for q_idx in range(1, questions_count + 1):
        correct_count = 0
        total_attempts = 0
        for student in students_list:
            att = attempts_map.get(student["id"])
            if att and att["completed"]:
                total_attempts += 1
                if att["answers"].get(q_idx):
                    correct_count += 1
        pct = (correct_count / total_attempts * 100) if total_attempts > 0 else 0
        # Agar imtihonda umuman hech kim qatnashmagan bo'lsa
        if total_attempts == 0:
            pct = 0
        q_stats.append({"q_num": q_idx, "correct_count": correct_count, "pct": pct})

    # To'g'ri topilgan foiz bo'yicha o'sish tartibida saralash (eng kichik foiz tepaga)
    q_stats_sorted = sorted(q_stats, key=lambda x: (x["pct"], x["correct_count"]))

    # Qoidalar:
    # 1. 0% yechilgan savollar bo'lsa - hammasi ro'yxatga kiradi.
    # 2. Aks holda - eng qiyin 3 ta savol.
    zero_pct_questions = [q for q in q_stats if q["pct"] == 0]
    hardest_list = []

    if len(zero_pct_questions) > 0:
        # Hamma 0% lik savollarni chiqaramiz
        hardest_list = zero_pct_questions
    else:
        # Eng qiyin 3 ta savolni olamiz
        hardest_list = q_stats_sorted[:3]

    # Tahlil natijalarini yozish
    curr_gap_row = gap_row + 1
    total_participants = len(active_students_rows)

    if total_participants == 0:
        ws2.cell(
            row=curr_gap_row, column=2, value="Imtihonda faol ishtirok etgan o'quvchilar mavjud emas."
        ).font = font_data
    else:
        for idx, stat in enumerate(hardest_list, 1):
            q_text = f"Savol {stat['q_num']}"
            ans_desc = f"{stat['correct_count']} ta o'quvchi to'g'ri topdi"
            pct_desc = f"{round(stat['pct'], 1)}%"

            line_text = f"{idx}. {q_text} ({ans_desc} - {pct_desc})"
            ws2.cell(row=curr_gap_row, column=2, value=line_text).font = font_bold
            ws2.cell(row=curr_gap_row, column=2).font = Font(name="Calibri", size=11, color="800000")
            curr_gap_row += 1

    # Ustunlar kengligini sozlash
    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        if col[0].column == 2:
            ws2.column_dimensions[col_letter].width = 30
        else:
            ws2.column_dimensions[col_letter].width = max(max_len + 2, 7)

    # Faylni xotiraga (memory stream) saqlash
    file_stream = io.BytesIO()
    wb.save(file_stream)
    return file_stream.getvalue()
